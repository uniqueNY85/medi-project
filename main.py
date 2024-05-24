# -*- coding: utf-8 -*-
"""
Created on Thu Aug 24 08:51:31 2023

@author: Jon Campbell
"""

import pandas as pd
from difflib import get_close_matches
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl import load_workbook
from datetime import datetime
import anvil.server
import anvil.stripe
import anvil.users
import uuid
from anvil.tables import app_tables

anvil.server.connect("server_WTQG27CU55CCGZCE4I4KFGN4-Q5FU3BNWECDMTS23")


#this is for the LoginForm and the SignUpForm

@anvil.server.callable
def authenticate_user(username, password, remember=False):
  #print(f"Authenticating: username={username}, password={password}, remember={remember}")  # Debugging print
  try:
    user = anvil.users.login_with_email(username, password)
    if remember:
      anvil.users.remember_login(duration_days=30)
    #print("Authentication successful")  # Debugging print
    return {'status': 'success'}
  except Exception as e:
    #print(f"Authentication failed: {e}")  # Debugging print
    return {'status': 'failure', 'message': str(e)}


@anvil.server.callable
def authenticate_google_user():
  try:
    user = anvil.users.login_with_google()
    return {'status': 'success'}
  except Exception as e:
    return {'status': 'failure', 'message': str(e)}

@anvil.server.callable
def set_guest_session():
    print("Setting guest session.")
    guest_id = str(uuid.uuid4())
    anvil.server.session['guest_id'] = guest_id
    return guest_id

@anvil.server.callable
def update_session_info(identifier, first_rpt_status_cd, second_identifier, first_beg_rpt_period_str, first_end_rpt_period_str, second_rpt_status_cd, second_beg_rpt_period_str, second_end_rpt_period_str, selected_year, provider_number, hha_name, address, user_email):
    user_row = app_tables.users.get(email=user_email)
    first_beg_rpt_period = datetime.strptime(first_beg_rpt_period_str, '%m/%d/%Y').date()
    first_end_rpt_period = datetime.strptime(first_end_rpt_period_str, '%m/%d/%Y').date()
    second_beg_rpt_period = datetime.strptime(second_beg_rpt_period_str, '%m/%d/%Y').date() if second_beg_rpt_period_str else None
    second_end_rpt_period = datetime.strptime(second_end_rpt_period_str, '%m/%d/%Y').date() if second_end_rpt_period_str else None
    if user_row is not None:
        new_row = app_tables.sessions_hha.add_row(
            fiscal_year=selected_year,
            provider_number=provider_number,
            hha_name=hha_name,
            hha_address=address,
            identifier=identifier,
            first_rpt_status_cd=first_rpt_status_cd,
            first_beg_rpt_period=first_beg_rpt_period, 
            first_end_rpt_period=first_end_rpt_period, 
            second_identifier=second_identifier,
            second_rpt_status_cd=second_rpt_status_cd,
            second_beg_rpt_period=second_beg_rpt_period, 
            second_end_rpt_period=second_end_rpt_period,
            operation_timestamp=datetime.now(),
            user_identifier=user_row,
            payment_status="Pending"
        )
        session_id = new_row.get_id()
        return session_id
    else:
        # Handle the case where user_row is None (user not found)
        print(f"User with email {user_email} not found.")


#the process starts here
#user input and search code
def direct_match_search(user_input, column, prvdr_info_df):
    """Directly filter out records that contain the user input."""
    lowered_input = user_input.lower()
    direct_matches = prvdr_info_df[prvdr_info_df[column].str.lower().str.contains(lowered_input)]
    return direct_matches

def search_provider(user_input, column, prvdr_info_df):
    """Search for a provider based on the user's input and specified column."""
    lowered_input = user_input.lower()
    lowered_data = prvdr_info_df[column].str.lower().tolist()
    matches = get_close_matches(lowered_input, lowered_data, n=50)
    return matches
@anvil.server.callable 
def get_user_input_and_search(user_input):
    try: 
        # Reading the CSV file to get the DataFrame
        prvdr_info_df = pd.read_csv('F:/Databases/Medicare Cost Reports/Home Health/HHA20-REPORTS/HHA20_PRVDR_ID_INFO.CSV')
    
        direct_matches_df = direct_match_search(user_input, 'HHA20_Name', prvdr_info_df)
    
        results_list = []
        
        if len(direct_matches_df) < 50:
            remaining_matches_needed = 50 - len(direct_matches_df)
            fuzzy_matches = search_provider(user_input, 'HHA20_Name', prvdr_info_df)
            fuzzy_matches_df = prvdr_info_df[prvdr_info_df['HHA20_Name'].str.lower().isin([x.lower() for x in fuzzy_matches])].head(remaining_matches_needed)
            combined_df = pd.concat([direct_matches_df, fuzzy_matches_df]).drop_duplicates().head(50)
        else:
            combined_df = direct_matches_df.head(50)
        
        counter = 1  # Initialize a counter variable
        if not combined_df.empty:
            for _, row in combined_df.iterrows():  # Replacing 'index' with '_'
                result_dict = {
                    'Selection_ID': counter,  # Use the counter variable here
                    'provider_number': row['PROVIDER_NUMBER'],
                    'hha_name': row['HHA20_Name'],
                    'address': f"{row['Street_Addr']}, {row['City']}, {row['State']} {row['Zip_Code']}"
                }
                results_list.append(result_dict)
                counter += 1  # Increment the counter
                
            return {"status": "Matches found", "results": results_list}
        else:
            return {"status": "No matches", "message": "No matches found. Please try again."}
            
    except Exception as e:
        return {"status": "Error", "message": str(e)}
    
@anvil.server.callable
def get_selected_data(data, user_selection_id):
  try:
    # Convert to integer and validate
    user_selection_id = int(user_selection_id)
    if user_selection_id > 0 and user_selection_id <= len(data):
      # Fetch the corresponding data
      selected_data = data[user_selection_id - 1]  # subtracting 1 because list indices start at 0
      return {"status": "Success", "selected_data": selected_data}
    else:
      return {"status": "Error", "message": f"Invalid Selection ID. Please enter a number between 1 and {len(data)}"}
  
  except ValueError:
    return {"status": "Error", "message": "Please enter a valid integer for Selection ID."}

@anvil.server.callable
def process_selected_data(provider_number): #I don't think this is being used
    try:
        # Check for available data years for this provider_number
        data_check_result = get_available_years(provider_number)

        if data_check_result["status"] == "success":
            available_years = data_check_result["available_years"]
            print(f"Available years for provider number {provider_number}: {available_years}")
            return {"status": "Success", "available_years": available_years, "message": "Data processed successfully"}
        else:
            return {"status": "Error", "message": "No available data years found"}

    except Exception as e:
        return {"status": "Error", "message": str(e)}

@anvil.server.callable
def get_available_years(provider_number):
    available_years = []
    year_identifiers = {}
    year_second_identifiers = {}  # New dictionary for possible second identifiers
    additional_data = {}  # Dictionary to store the additional information for the first identifier
    additional_second_data = {} # New dictionary for additional information for the second identifier
    try:
        # File Paths
        FY2021_file_path = "F:/Databases/Medicare Cost Reports/Home Health/HHA20FY2021/HHA20_2021_RPT.CSV"
        FY2022_file_path = "F:/Databases/Medicare Cost Reports/Home Health/HHA20FY2022/HHA20_2022_RPT.CSV"
        FY2023_file_path = "F:/Databases/Medicare Cost Reports/Home Health/HHA20FY2023/HHA20_2023_RPT.CSV"
        # Load CSVs into DataFrames
        rpt_df_2021 = pd.read_csv(FY2021_file_path, header=None)
        rpt_df_2022 = pd.read_csv(FY2022_file_path, header=None)
        rpt_df_2023 = pd.read_csv(FY2023_file_path, header=None)
        # Function to extract additional data from a given row
        def extract_additional_data(row):
            # If row is a DataFrame with a single row, extract the Series
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            return {
                'rpt_status_code': row[4],  # Column E
                'beg_rpt_period': row[5],   # Column F
                'end_rpt_period': row[6]    # Column G
            }   

        # Check FY2021
        matching_rows_2021 = rpt_df_2021[rpt_df_2021[2] == provider_number]
        print(f"Matching rows in FY2021 for provider {provider_number}: {len(matching_rows_2021)}")
        if not matching_rows_2021.empty:
            available_years.append("FY2021")
            year_identifiers["FY2021"] = matching_rows_2021.iloc[0, 0]
            additional_data["FY2021"] = extract_additional_data(matching_rows_2021.iloc[0])
            if len(matching_rows_2021) > 1:  # If more than one identifier exists
                year_second_identifiers["FY2021"] = matching_rows_2021.iloc[1, 0]
                additional_second_data["FY2021"] = extract_additional_data(matching_rows_2021.iloc[1])

        # Check FY2022
        matching_rows_2022 = rpt_df_2022[rpt_df_2022[2] == provider_number]
        print(f"Matching rows in FY2022 for provider {provider_number}: {len(matching_rows_2022)}")
        if not matching_rows_2022.empty:
            available_years.append("FY2022")
            year_identifiers["FY2022"] = matching_rows_2022.iloc[0, 0]
            additional_data["FY2022"] = extract_additional_data(matching_rows_2022)
            if len(matching_rows_2022) > 1:
                year_second_identifiers["FY2022"] = matching_rows_2022.iloc[1, 0]
                additional_second_data["FY2022"] = extract_additional_data(matching_rows_2022.iloc[1]) 
                
        # Check FY2023
        matching_rows_2023 = rpt_df_2023[rpt_df_2023[2] == provider_number]
        print(f"Matching rows in FY2023 for provider {provider_number}: {len(matching_rows_2023)}")
        if not matching_rows_2023.empty:
            available_years.append("FY2023")
            year_identifiers["FY2023"] = matching_rows_2023.iloc[0, 0]
            additional_data["FY2023"] = extract_additional_data(matching_rows_2023)
            if len(matching_rows_2023) > 1:
                year_second_identifiers["FY2023"] = matching_rows_2023.iloc[1, 0]
                additional_second_data["FY2023"] = extract_additional_data(matching_rows_2023.iloc[1])
        print(f"Available years: {available_years}")
        print(f"Year Identifiers: {year_identifiers}")
        print(f"Year Second Identifiers: {year_second_identifiers}")
        print(f"Additional Data: {additional_data}")
        print(f"Additional Second Data: {additional_second_data}")
        if available_years:
            return {
                "status": "success",
                "available_years": available_years,
                "identifiers": year_identifiers,
                "second_identifiers": year_second_identifiers,
                "additional_data": additional_data,  # Additional data for the first identifier
                "additional_second_data": additional_second_data  # Additional data for the second identifier
            }
        else:
            return {"status": "failure", "message": "No available data years found"}
    except Exception as e:
        print(f"An exception occurred: {e}")
        return {"status": "error", "message": str(e)}

@anvil.server.callable
def charge_user(token, email, selected_year, identifier, session_id, second_identifier): 
    session_row = None  # Initialize session_row to None
    try:
        stripe_customer = anvil.stripe.new_customer(email, token)
        charge = stripe_customer.charge(amount=5000, currency="usd")
        # Fetch the session row based on the session_id
        if session_id:
            session_row = app_tables.sessions_hha.get_by_id(session_id)
        if charge['status'] == 'succeeded':
            receipt_url = charge['receipt_url']  # Capture the receipt URL
            if session_row:
                session_row['payment_status'] = 'Succeeded'
                session_row['receipt_url'] = receipt_url  # Store the receipt URL in the session row
            try:
                process_user_rpt(selected_year, identifier, session_id, second_identifier)
                if session_row:
                    session_row['report_gen_status'] = 'Succeeded'  # Update report__gen_status on successful report generation
                return {'payment_status': 'succeeded', 'report_gen_status': 'succeeded'}
            except Exception as e_inner:
                print(f"An error occurred during report processing: {str(e_inner)}")
                if session_row:
                    session_row['report_gen_status'] = 'Report Error'
                return {'payment_status': 'succeeded', 'report_gen_status': 'report_error'}
        else:
            if session_row:
                session_row['payment_status'] = 'Charge Failed'
            return 'charge_failed'
    except Exception as e_outer:
        print(f"An error occurred: {str(e_outer)}")
        session_row = app_tables.sessions_hha.get_by_id(session_id)
        if session_row:
            session_row['payment_status'] = 'Charge Error'
        return 'charge_error'

def strip_leading_zeros(val):
    # If the entire string is "0", return it as is
    if val == "0":
        return val
    # Check if the value is purely numeric
    elif val.isnumeric():
        return val.lstrip('0')
    # If alphanumeric or other cases
    else:
        return val

def fetch_nmrc_data(nmrc_file_path, identifier):
    # Read the CSV file into a DataFrame with specified column names and types
    col_names = ["Identifier", "WorksheetNumber", "LineNumber", "ColumnNumber", "Value"]
    hha_nmrc_df = pd.read_csv(nmrc_file_path, names=col_names, header=None, dtype={'ColumnNumber': str})
    # Filter hha_df based on the identifier
    hha_nmrc_df_filtered = hha_nmrc_df[hha_nmrc_df['Identifier'] == identifier].copy()
    # Apply the strip_leading_zeros function to the ColumnNumber
    hha_nmrc_df_filtered['ColumnNumber'] = hha_nmrc_df_filtered['ColumnNumber'].apply(strip_leading_zeros)
    # Replace empty strings with "0"
    hha_nmrc_df_filtered['ColumnNumber'].replace('', '0', inplace=True)
    return hha_nmrc_df_filtered

def fetch_alpha_data(alpha_file_path, identifier):
    # Read the CSV file into a DataFrame with specified column names and types
    col_names = ["Identifier", "WorksheetNumber", "LineNumber", "ColumnNumber", "Value"]
    hha_alpha_df = pd.read_csv(alpha_file_path, names=col_names, header=None, dtype={'ColumnNumber': str})
    # Filter hha_df based on the identifier
    hha_alpha_df_filtered = hha_alpha_df[hha_alpha_df['Identifier'] == identifier].copy()  
    # Apply the strip_leading_zeros function to the ColumnNumber
    hha_alpha_df_filtered['ColumnNumber'] = hha_alpha_df_filtered['ColumnNumber'].apply(strip_leading_zeros)
    # Replace empty strings with "0"
    hha_alpha_df_filtered['ColumnNumber'].replace('', '0', inplace=True)
    return hha_alpha_df_filtered

#this will be called if successful payment in charge_user after if charge['status'] == 'succeeded':
def process_user_rpt(selected_year, identifier, session_id, second_identifier=None): 
    # Fetch session info from Anvil's data table
    session_row = app_tables.sessions_hha.get_by_id(session_id)
    hha_name = session_row['hha_name']
    hha_address = session_row['hha_address']
    provider_number = session_row['provider_number']
    rpt_status_code = session_row['first_rpt_status_cd']
    beg_rpt_period = session_row['first_beg_rpt_period']
    end_rpt_period = session_row['first_end_rpt_period']
    # Template Path and Destination Folder
    template_path = 'F:/Databases/Medicare Cost Reports/Home Health/Python Code/HHA MC Cost Report Worksheets.xlsx'
    destination_folder = 'F:/Databases/Medicare Cost Reports/Home Health/Web App Output/'
    #format selected_year for the unique_file_name
    cleaned_year = selected_year.replace("FY", "")
    # Generating a timestamp for the unique file name
    timestamp_str = datetime.now().strftime('%Y%m%d%H%M%S')
    # Forming the unique file name
    unique_file_name = f"{selected_year}_{identifier}_medicare_cost_report_data_{timestamp_str}.xlsx"
    # Load the existing workbook
    destination_wb = load_workbook(filename=template_path)
    # Form the complete destination path with the unique file name
    destination_path = f"{destination_folder}/{unique_file_name}"
    # Save the workbook to create the file on disk to allow appending nmrc and alpha worksheets
    destination_wb.save(destination_path)
    # Fetch nmrc and alpha data
    nmrc_file_path = f"F:/Databases/Medicare Cost Reports/Home Health/HHA20{selected_year}/HHA20_{cleaned_year}_NMRC.CSV"
    nmrc_data = fetch_nmrc_data(nmrc_file_path, identifier)
    alpha_file_path = f"F:/Databases/Medicare Cost Reports/Home Health/HHA20{selected_year}/HHA20_{cleaned_year}_ALPHA.CSV"
    alpha_data = fetch_alpha_data(alpha_file_path, identifier)
    # Create a new sheet for nmrc_data
    ws_nmrc = destination_wb.create_sheet("Numeric Data")
    for row in dataframe_to_rows(nmrc_data, index=False, header=True):
        ws_nmrc.append(row)
    # Create a new sheet for alpha_data
    ws_alpha = destination_wb.create_sheet("Alpha Data")
    for row in dataframe_to_rows(alpha_data, index=False, header=True):
        ws_alpha.append(row)
    ws_info = destination_wb.create_sheet("Info")
    # Move the "Info" sheet to the first position
    # Find the current index (position) of the "Info" sheet
    info_index = destination_wb.sheetnames.index("Info")
    # Calculate the offset needed to move the "Info" sheet to the first position
    offset = -info_index
    # Move the "Info" sheet to the first position using the calculated offset
    destination_wb.move_sheet(ws_info, offset=offset)
    #Create a dictionary for the status codes
    status_code_mapping = {
        1: "As submitted",
        2: "Settled w/o audit",
        3: "Settled w/ audit",
        4: "Reopened",
        5: "Amended"
    }
    info_data = [
        ["Fiscal Year:", selected_year],
        ["Report Period:", f"{beg_rpt_period} - {end_rpt_period}"],  
        ["Report Status:",rpt_status_code],       
        ["HHA Name:", hha_name],
        ["Provider Number:", provider_number],
        ["Medicare Report Identifier:", identifier],
        ["Full Address:", hha_address]
    ]
    # Replace the numerical status code with its textual meaning
    info_data[2][1] = status_code_mapping[rpt_status_code]
    for row in info_data:
        ws_info.append(row)
    # Formatting the "Info" worksheet
    for row in ws_info.iter_rows():
        for cell in row:
            cell.font = Font(bold=True) if cell.column == 1 else Font(bold=False)
      
     
    # Define the mapping from worksheet names to worksheet numbers
    worksheet_map = {
        'Worksheet_S-3_Pts I-III': 'S300000',
        'Worksheet_S-3_Pt IV': 'S300004',
        'Worksheet_S-3_Pt V': 'S300005',
        'Worksheet_A': 'A000000',
        'Worksheet_B': 'B000000',
        'Worksheet_C': 'C000000',
        'Worksheet_F': 'F000000',
        'Worksheet_F-1': 'F100000'
    }
    #fetch the values from nmrc and alpha data frames to prepare to populate target worksheets.
    worksheet_name_map = {v: k for k, v in worksheet_map.items()}
    # Define the data types for the DataFrame
    data_types = {
        'WorksheetNumber': 'object',  # strings
        'LineNumber': 'int',         # Line numbers are integers
        'ColumnNumber': 'object',    # Column numbers might be alphanumeric
        'Value': 'object'             # has to be object 
    }

    # Initialize the DataFrame with specified data types
    fetched_rows_df = pd.DataFrame(columns=data_types.keys()).astype(data_types)
    desired_worksheets = {'S300000', 'S300004', 'S300005', 'A000000', 'B000000', 'C000000', 'F000000', 'F100000'}

    # Loop through both source DataFrames
    for source_data in [nmrc_data, alpha_data]:
        for _, row in source_data.iterrows():
            worksheet_number = row['WorksheetNumber']
            worksheet_name = worksheet_name_map.get(worksheet_number, "Unknown Worksheet")
            column_number = str(row["ColumnNumber"])  # Ensure column number is a string
            line_number = row["LineNumber"]

            if worksheet_number not in desired_worksheets:
                continue  # Skip this iteration of the loop

            if source_data is nmrc_data:
                column_number_for_mask = column_number  # Already a string
            else:
                try:
                    column_number_for_mask = str(int(column_number))  # Convert to integer then back to string
                except ValueError:  # If conversion fails
                    column_number_for_mask = column_number  # Use the original string

            mask = (line_number == int(line_number)) & (column_number == column_number_for_mask)
            if mask:
                new_row = {
                    'WorksheetNumber': worksheet_number,
                    'LineNumber': line_number,
                    'ColumnNumber': column_number,
                    'Value': row['Value']  # Keep the original data type
                }
                new_row_df = pd.DataFrame([new_row])  # Convert the new row to a DataFrame

            # Ensure new_row_df matches the data types defined
            new_row_df = new_row_df.astype(data_types)

            # Concatenate the DataFrames
            fetched_rows_df = pd.concat([fetched_rows_df, new_row_df], ignore_index=True)
           # else : #debug print
              #  print(f"Row with LineNumber {row['LineNumber']} and ColumnNumber {column_number} did not pass the mask condition.")
             
    destination_wb.save(destination_path)

#this is where populating target worksheets occurs
    line_ranges_map = {}
    col_ranges_map = {}
    populating_df = pd.DataFrame(columns=['WorksheetName', 'ExcelRow', 'ExcelColumn', 'Value'])
    # Map the columns from fetched_rows_df to populating_df
    populating_df['WorksheetName'] = fetched_rows_df['WorksheetNumber'].map(worksheet_name_map)
    populating_df['ExcelRow'] = fetched_rows_df['LineNumber']
    populating_df['ExcelColumn'] = fetched_rows_df['ColumnNumber']
    populating_df['Value'] = fetched_rows_df['Value']

    # Define starting points for line and column numbers in each worksheet
    line_start_map = {
        'Worksheet_S-3_Pts I-III': [8, 23, 26, 48],
        'Worksheet_S-3_Pt IV': [7],
        'Worksheet_S-3_Pt V': [8],
        'Worksheet_A': [5], 
        'Worksheet_B': [5, 51], 
        'Worksheet_C': [8, 24],
        'Worksheet_F': [6], 
        'Worksheet_F-1': [6, 11]
    }
    col_start_map = {
        'Worksheet_S-3_Pts I-III': [3, 3, 2, 9],
        'Worksheet_S-3_Pt IV': [3],
        'Worksheet_S-3_Pt V': [4],
        'Worksheet_A': [2], 
        'Worksheet_B': [3, 3], 
        'Worksheet_C': [7, 4], 
        'Worksheet_F': [5], 
        'Worksheet_F-1': [3, 2]
    }
    #Define the ending points for line and column numbers in each worksheet
    line_end_map = {
        'Worksheet_S-3_Pts I-III': [20, 23, 44, 55],
        'Worksheet_S-3_Pt IV': [24],
        'Worksheet_S-3_Pt V': [36],
        'Worksheet_A': [46], 
        'Worksheet_B': [46, 92], 
        'Worksheet_C': [17, 27],
        'Worksheet_F': [62], 
        'Worksheet_F-1': [8, 42]
    }
    col_end_map = {
        'Worksheet_S-3_Pts I-III': [10, 3, 9, 9],
        'Worksheet_S-3_Pt IV': [7],
        'Worksheet_S-3_Pt V': [8],
        'Worksheet_A': [12], 
        'Worksheet_B': [9, 9], 
        'Worksheet_C': [12, 12], 
        'Worksheet_F': [5], 
        'Worksheet_F-1': [6, 6]
    }
    #mapping to hold Excel row indices for column numbers in each worksheet
    col_row_map = {
        'Worksheet_S-3_Pts I-III': [7, 22, 25, 47],
        'Worksheet_S-3_Pt IV': [6],
        'Worksheet_S-3_Pt V': [5],
        'Worksheet_A': [3],  # corresponding to Excel row 3
        'Worksheet_B': [3, 49],  # corresponding to Excel rows 3 and 49
        'Worksheet_C': [7, 23],
        'Worksheet_F': [3],
        'Worksheet_F-1': [5,10]
    }
    with pd.ExcelFile(destination_path) as xls:     
        # Read ranges for Worksheets
        
        df_S3_0 = pd.read_excel(xls, 'Worksheet_S-3_Pts I-III', header=None)
        line_ranges_map['Worksheet_S-3_Pts I-III'] = [
            [x for x in df_S3_0.iloc[7:20, 0].tolist() if isinstance(x, int)],  # Range 1: A8:A23
            [x for x in df_S3_0.iloc[22:23, 0].tolist() if isinstance(x, int)],  # Range 2: A23
            [x for x in df_S3_0.iloc[25:44, 0].tolist() if isinstance(x, int)],  # Range 3: A26:A44
            [x for x in df_S3_0.iloc[47:55, 0].tolist() if isinstance(x, int)]   # Range 4: A48:A55
        ]
        col_ranges_map['Worksheet_S-3_Pts I-III'] = [
            [str(x) for x in df_S3_0.iloc[6, 2:10].tolist()],   # Range 1: C7:J7
            [str(x) for x in df_S3_0.iloc[21, 2:3].tolist()],   # Range 2: C22
            [str(x) for x in df_S3_0.iloc[24, 1:9].tolist()],   # Range 3: B25:I25
            [str(x) for x in df_S3_0.iloc[46, 8:9].tolist()]    # Range 4: I47
        ]   

# Read the Excel file for Worksheet_S-3_Pt IV
        df_S3_4 = pd.read_excel(xls, 'Worksheet_S-3_Pt IV', header=None)
        line_ranges_map['Worksheet_S-3_Pt IV'] = [
            [x for x in df_S3_4.iloc[6:24, 0].tolist() if isinstance(x, int)],  # Range 1: A7:A24
        ]
        col_ranges_map['Worksheet_S-3_Pt IV'] = [
            [str(x) for x in df_S3_4.iloc[5, 2:7].tolist()],   # Range 1: C6:G6
        ]

# Read the Excel file for Worksheet_S-3_Pt V
        df_S3_5 = pd.read_excel(xls, 'Worksheet_S-3_Pt V', header=None)
        line_ranges_map['Worksheet_S-3_Pt V'] = [
            [x for x in df_S3_5.iloc[7:36, 0].tolist() if isinstance(x, int)],  # Range 1: A8:A36
        ]
        col_ranges_map['Worksheet_S-3_Pt V'] = [
            [str(x) for x in df_S3_5.iloc[4, 3:8].tolist()],   # Range 1: D5:I5
        ]
                
        df_A = pd.read_excel(xls, 'Worksheet_A', header=None)
        line_ranges_map['Worksheet_A'] = [[x for x in df_A.iloc[4:46, 0].tolist() if isinstance(x, int)]]  #A5:A46
        col_ranges_map['Worksheet_A'] = [
            [str(x) for x in df_A.iloc[2, 1:12].tolist()]  #B3:L3
        ]
        
        df_B = pd.read_excel(xls, 'Worksheet_B', header=None)
        line_ranges_map['Worksheet_B'] = [
            [x for x in df_B.iloc[4:46, 0].tolist() if isinstance(x, int)],  # for the first range A5:A46
            [x for x in df_B.iloc[50:92, 0].tolist() if isinstance(x, int)]  # for the second range A51:A92
        ]
        col_ranges_map['Worksheet_B'] = [
            [str(x) for x in df_B.iloc[2, 2:9].tolist()], #C3:I3
            [str(x) for x in df_B.iloc[48, 2:9].tolist()]  #C49:I49
        ]  
        df_C = pd.read_excel(xls, 'Worksheet_C', header=None)
        line_ranges_map['Worksheet_C'] = [
            [x for x in df_C.iloc[7:17, 0].tolist() if isinstance(x, int)],  #A8:A17
            [x for x in df_C.iloc[23:27, 0].tolist() if isinstance(x, int)]   #A24:A27
        ]  
        col_ranges_map['Worksheet_C'] = [
            [str(x) for x in df_C.iloc[6, 7:12].tolist()],   #H7:L7
            [str(x) for x in df_C.iloc[22, 3:12].tolist()]   #D23:L23
        ]
    
        df_F = pd.read_excel(xls, 'Worksheet_F', header=None)
        line_ranges_map['Worksheet_F'] = [[x for x in df_F.iloc[5:62, 0].tolist() if isinstance(x, int)]]  #A6:A62
        col_ranges_map['Worksheet_F'] = [
            [str(x) for x in df_F.iloc[2, 4:5].tolist()]  #E3
        ]
    
        df_F1 = pd.read_excel(xls, 'Worksheet_F-1', header=None)
        line_ranges_map['Worksheet_F-1'] = [
            [x for x in df_F1.iloc[5:8, 0].tolist() if isinstance(x, int)], # A6:A8
            [x for x in df_F1.iloc[10:42, 0].tolist() if isinstance(x, int)] # A11:A42
        ]  # Two different ranges
        col_ranges_map['Worksheet_F-1'] = [
            [str(x) for x in df_F1.iloc[4, 2:6].tolist()],  # C5:F5
            [str(x) for x in df_F1.iloc[9, 1:6].tolist()]   # B10:F10
        ]       
    # Step 1: Loop through each target worksheet
    with pd.ExcelWriter(destination_path, engine='openpyxl', mode='a') as writer:
        #print("Entered Excel Writer block.")
        #initialize a "master" dictionary to hold line_to_row and col_to_column mappings for each range in each worksheet
        master_line_to_row = {}
        master_col_to_column = {}
        for worksheet_name in line_ranges_map.keys():
           # print(f"Working on worksheet: {worksheet_name}.")
            ws = writer.book[worksheet_name]
            # Initialize an empty mapping dictionary
            line_to_row = {}
            col_to_column = {}
            # Loop through the list of starting and ending points for line and column
            for idx, (start_row, end_row, start_col, end_col) in enumerate(zip(
                line_start_map[worksheet_name], line_end_map[worksheet_name], 
                col_start_map[worksheet_name], col_end_map[worksheet_name])):
                #print(f"Line to Row Mapping: {line_to_row}")  
                #print(f"Col to Column Mapping: {col_to_column}")
                # Create a unique key for each range within each worksheet
                unique_range_key = f"{worksheet_name}_{idx}"
                # Store line_to_row and col_to_column mapping for each unique range
                master_line_to_row[unique_range_key] = map_excel_rows_to_lines(ws, start_row, end_row)
                #print(f"Current line_to_row mapping for {unique_range_key}: {master_line_to_row[unique_range_key]}")
                # Fetch the row index to look for columns based on the current range
                col_row_index = col_row_map.get(worksheet_name, [3])[idx]  # default to 3 if not found
                master_col_to_column[unique_range_key] = map_excel_cols_to_columns(ws, start_col, end_col, col_row_index)
                #print(f"Current col_to_column mapping for {unique_range_key}: {master_col_to_column[unique_range_key]}")
                # Print the types of the column values stored within the master_col_to_column dictionary for the current unique_range_key
                #for col_key, col_value in master_col_to_column[unique_range_key].items():
                    #print(f"Column Key: {col_key}, Type: {type(col_key)}, Column Value: {col_value}, Type: {type(col_value)}")
            # Continue with the code that populates the Excel sheet
            df_filtered = populating_df[populating_df['WorksheetName'] == worksheet_name]        
            for _, row in df_filtered.iterrows():
                #print(f"DataFrame Row: {row}")
                line = row['ExcelRow']
                col = row['ExcelColumn']
                value = row['Value']
                #print(f"Type and value of line: {type(line)}, {line}") 
                #print(f"Type and value of col: {type(col)}, {col}")
                #print(f"Debug - About to look up col: {col} of type: {type(col)}")
                # Determine which unique range key to use 
                unique_range_key = determine_unique_range_key(line, col, line_ranges_map, col_ranges_map, worksheet_name)
                # Check if a unique_range_key was found
                if unique_range_key is None:
                    #print(f"Skipping: No unique range key found for Line {line}, Column {col}")
                    continue  # Skip to next iteration
                if unique_range_key:
                    excel_row = master_line_to_row.get(unique_range_key, {}).get(line)
                    excel_col = master_col_to_column.get(unique_range_key, {}).get(col)
                if excel_row is not None and excel_col is not None:
                    #print(f"Writing to Excel: Row {excel_row}, Column {excel_col}, Value {value}")
                    ws.cell(row=excel_row, column=excel_col, value=value)
                #else:
                    #print(f"Skipping: Excel Row or Excel Col is None. Excel Row: {excel_row}, Excel Col: {excel_col}")
    # Save the workbook again after adding the sheets, populating, and formatting
        destination_wb.save(destination_path)  
        
        if second_identifier:
        # You might want to log or inform that the second report processing is starting
            print(f"Processing second report for session_id {session_id}")
            process_second_report(selected_year, second_identifier, session_id)
            
def map_excel_rows_to_lines(ws, start_row, end_row):
    #print(f"Mapping rows to lines for start_row: {start_row}, end_row: {end_row}")
    line_to_row = {}
    for row in range(start_row, end_row + 1):
        line_number = ws.cell(row=row, column=1).value  # Assuming line numbers are in column 'A'
        if isinstance(line_number, int):  # Only map integers
            line_to_row[line_number] = row
            #print(f"Type and value of line_number: {type(line_number)}, {line_number}")
            #print(f"Type and value of row: {type(row)}, {row}")
        #else:
            #print(f"Skipping non-integer line_number: {line_number} of type {type(line_number)}")
    return line_to_row
def map_excel_cols_to_columns(ws, start_col, end_col, col_row_index):
    #print(f"Mapping columns to columns for start_col: {start_col}, end_col: {end_col}")
    col_to_column = {}
    for col in range(start_col, end_col + 1):
        col_number = ws.cell(row=col_row_index, column=col).value  # This is now dynamic
        if isinstance(col_number, int): 
            col_number = str(col_number)
        #print(f"Type and value of col_number: {type(col_number)}, {col_number}")
        #print(f"Type and value of col: {type(col)}, {col}")
        col_to_column[col_number] = col
    return col_to_column

def determine_unique_range_key(line, col, line_ranges_map, col_ranges_map, worksheet_name):
    """
    Determine the unique range key for a given line and column combination.
    """
    # Print the master dictionary ranges
    #print("---- Master Dictionary Ranges ----")
    #print("Line Ranges Map:", line_ranges_map)
    #print("Col Ranges Map:", col_ranges_map)
    #print("----------------------------------") 

    line_ranges = line_ranges_map[worksheet_name]
    for idx, (line_range, col_range) in enumerate(zip(line_ranges, col_ranges_map[worksheet_name])):
        # Print the ranges being checked
        #print(f"Checking line: {line}, col: {col} against line_range: {line_range} and col_range: {col_range}")
        if line in line_range and col in col_range:
            return f"{worksheet_name}_{idx}"  # return the unique range key
    #print(f"Skipping: No unique range key found for Line {line}, Column {col}")
    return None  # If no unique range key is found

#for the below function, will need to grab the second status and report periods from the session table
#like so rpt_status_code = session_row['second_rpt_status_cd']  [second_beg_rpt_period]   [second_end_rpt_period]
def process_second_report(selected_year, identifier, session_id):
    # Fetch session info from Anvil's data table
    session_row = app_tables.sessions_hha.get_by_id(session_id)
    hha_name = session_row['hha_name']
    hha_address = session_row['hha_address']
    provider_number = session_row['provider_number']
    rpt_status_code = session_row['second_rpt_status_cd']
    beg_rpt_period = session_row['second_beg_rpt_period']
    end_rpt_period = session_row['second_end_rpt_period']
    # Template Path and Destination Folder
    template_path = 'F:/Databases/Medicare Cost Reports/Home Health/Python Code/HHA MC Cost Report Worksheets.xlsx'
    destination_folder = 'F:/Databases/Medicare Cost Reports/Home Health/Web App Output/'
    #format selected_year for the unique_file_name
    cleaned_year = selected_year.replace("FY", "")
    # Generating a timestamp for the unique file name
    timestamp_str = datetime.now().strftime('%Y%m%d%H%M%S')
    # Forming the unique file name
    unique_file_name = f"{selected_year}_{identifier}_medicare_cost_report_data_{timestamp_str}.xlsx"
    # Load the existing workbook
    destination_wb = load_workbook(filename=template_path)
    # Form the complete destination path with the unique file name
    destination_path = f"{destination_folder}/{unique_file_name}"
    # Save the workbook to create the file on disk to allow appending nmrc and alpha worksheets
    destination_wb.save(destination_path)
    # Fetch nmrc and alpha data
    nmrc_file_path = f"F:/Databases/Medicare Cost Reports/Home Health/HHA20{selected_year}/HHA20_{cleaned_year}_NMRC.CSV"
    nmrc_data = fetch_nmrc_data(nmrc_file_path, identifier)
    alpha_file_path = f"F:/Databases/Medicare Cost Reports/Home Health/HHA20{selected_year}/HHA20_{cleaned_year}_ALPHA.CSV"
    alpha_data = fetch_alpha_data(alpha_file_path, identifier)
    # Create a new sheet for nmrc_data
    ws_nmrc = destination_wb.create_sheet("Numeric Data")
    for row in dataframe_to_rows(nmrc_data, index=False, header=True):
        ws_nmrc.append(row)
    # Create a new sheet for alpha_data
    ws_alpha = destination_wb.create_sheet("Alpha Data")
    for row in dataframe_to_rows(alpha_data, index=False, header=True):
        ws_alpha.append(row)
    #create and populate the "Info" worksheet
    ws_info = destination_wb.create_sheet("Info")
    # Move the "Info" sheet to the first position
    # Find the current index (position) of the "Info" sheet
    info_index = destination_wb.sheetnames.index("Info")
    # Calculate the offset needed to move the "Info" sheet to the first position
    offset = -info_index
    # Move the "Info" sheet to the first position using the calculated offset
    destination_wb.move_sheet(ws_info, offset=offset)
    #Create a dictionary for the status codes
    status_code_mapping = {
        1: "As submitted",
        2: "Settled w/o audit",
        3: "Settled w/ audit",
        4: "Reopened",
        5: "Amended"
    }
    info_data = [
        ["Fiscal Year:", selected_year],
        ["Report Period:", f"{beg_rpt_period} - {end_rpt_period}"],  
        ["Report Status:",rpt_status_code],       
        ["HHA Name:", hha_name],
        ["Provider Number:", provider_number],
        ["Medicare Report Identifier:", identifier],
        ["Full Address:", hha_address]
    ]
    # Replace the numerical status code with its textual meaning
    info_data[2][1] = status_code_mapping[rpt_status_code]
    
    for row in info_data:
        ws_info.append(row)
    # Formatting the "Info" worksheet
    for row in ws_info.iter_rows():
        for cell in row:
            cell.font = Font(bold=True) if cell.column == 1 else Font(bold=False)
   
     
    # Define the mapping from worksheet names to worksheet numbers
    worksheet_map = {
        'Worksheet_S-3_Pts I-III': 'S300000',
        'Worksheet_S-3_Pt IV': 'S300004',
        'Worksheet_S-3_Pt V': 'S300005',
        'Worksheet_A': 'A000000',
        'Worksheet_B': 'B000000',
        'Worksheet_C': 'C000000',
        'Worksheet_F': 'F000000',
        'Worksheet_F-1': 'F100000'
    }
    #fetch the values from nmrc and alpha data frames to prepare to populate target worksheets.
    worksheet_name_map = {v: k for k, v in worksheet_map.items()}
    # Define the data types for the DataFrame
    data_types = {
        'WorksheetNumber': 'object',  # strings
        'LineNumber': 'int',         # Line numbers are integers
        'ColumnNumber': 'object',    # Column numbers might be alphanumeric
        'Value': 'object'             # has to be object 
    }

    # Initialize the DataFrame with specified data types
    fetched_rows_df = pd.DataFrame(columns=data_types.keys()).astype(data_types)
    desired_worksheets = {'S300000', 'S300004', 'S300005', 'A000000', 'B000000', 'C000000', 'F000000', 'F100000'}

    # Loop through both source DataFrames
    for source_data in [nmrc_data, alpha_data]:
        for _, row in source_data.iterrows():
            worksheet_number = row['WorksheetNumber']
            worksheet_name = worksheet_name_map.get(worksheet_number, "Unknown Worksheet")
            column_number = str(row["ColumnNumber"])  # Ensure column number is a string
            line_number = row["LineNumber"]

            if worksheet_number not in desired_worksheets:
                continue  # Skip this iteration of the loop

            if source_data is nmrc_data:
                column_number_for_mask = column_number  # Already a string
            else:
                try:
                    column_number_for_mask = str(int(column_number))  # Convert to integer then back to string
                except ValueError:  # If conversion fails
                    column_number_for_mask = column_number  # Use the original string

            mask = (line_number == int(line_number)) & (column_number == column_number_for_mask)
            if mask:
                new_row = {
                    'WorksheetNumber': worksheet_number,
                    'LineNumber': line_number,
                    'ColumnNumber': column_number,
                    'Value': row['Value']  # Keep the original data type
                }
                new_row_df = pd.DataFrame([new_row])  # Convert the new row to a DataFrame

            # Ensure new_row_df matches the data types defined
            new_row_df = new_row_df.astype(data_types)

            # Concatenate the DataFrames
            fetched_rows_df = pd.concat([fetched_rows_df, new_row_df], ignore_index=True)
           # else : #debug print
              #  print(f"Row with LineNumber {row['LineNumber']} and ColumnNumber {column_number} did not pass the mask condition.")
             
    destination_wb.save(destination_path)

#this is where populating target worksheets occurs
    line_ranges_map = {}
    col_ranges_map = {}
    populating_df = pd.DataFrame(columns=['WorksheetName', 'ExcelRow', 'ExcelColumn', 'Value'])
    # Map the columns from fetched_rows_df to populating_df
    populating_df['WorksheetName'] = fetched_rows_df['WorksheetNumber'].map(worksheet_name_map)
    populating_df['ExcelRow'] = fetched_rows_df['LineNumber']
    populating_df['ExcelColumn'] = fetched_rows_df['ColumnNumber']
    populating_df['Value'] = fetched_rows_df['Value']

    # Define starting points for line and column numbers in each worksheet
    line_start_map = {
        'Worksheet_S-3_Pts I-III': [8, 23, 26, 48],
        'Worksheet_S-3_Pt IV': [7],
        'Worksheet_S-3_Pt V': [8],
        'Worksheet_A': [5], 
        'Worksheet_B': [5, 51], 
        'Worksheet_C': [8, 24],
        'Worksheet_F': [6], 
        'Worksheet_F-1': [6, 11]
    }
    col_start_map = {
        'Worksheet_S-3_Pts I-III': [3, 3, 2, 9],
        'Worksheet_S-3_Pt IV': [3],
        'Worksheet_S-3_Pt V': [4],
        'Worksheet_A': [2], 
        'Worksheet_B': [3, 3], 
        'Worksheet_C': [7, 4], 
        'Worksheet_F': [5], 
        'Worksheet_F-1': [3, 2]
    }
    #Define the ending points for line and column numbers in each worksheet
    line_end_map = {
        'Worksheet_S-3_Pts I-III': [20, 23, 44, 55],
        'Worksheet_S-3_Pt IV': [24],
        'Worksheet_S-3_Pt V': [36],
        'Worksheet_A': [46], 
        'Worksheet_B': [46, 92], 
        'Worksheet_C': [17, 27],
        'Worksheet_F': [62], 
        'Worksheet_F-1': [8, 42]
    }
    col_end_map = {
        'Worksheet_S-3_Pts I-III': [10, 3, 9, 9],
        'Worksheet_S-3_Pt IV': [7],
        'Worksheet_S-3_Pt V': [8],
        'Worksheet_A': [12], 
        'Worksheet_B': [9, 9], 
        'Worksheet_C': [12, 12], 
        'Worksheet_F': [5], 
        'Worksheet_F-1': [6, 6]
    }
    #mapping to hold Excel row indices for column numbers in each worksheet
    col_row_map = {
        'Worksheet_S-3_Pts I-III': [7, 22, 25, 47],
        'Worksheet_S-3_Pt IV': [6],
        'Worksheet_S-3_Pt V': [5],
        'Worksheet_A': [3],  # corresponding to Excel row 3
        'Worksheet_B': [3, 49],  # corresponding to Excel rows 3 and 49
        'Worksheet_C': [7, 23],
        'Worksheet_F': [3],
        'Worksheet_F-1': [5,10]
    }
    with pd.ExcelFile(destination_path) as xls:     
        # Read ranges for Worksheets
        
        df_S3_0 = pd.read_excel(xls, 'Worksheet_S-3_Pts I-III', header=None)
        line_ranges_map['Worksheet_S-3_Pts I-III'] = [
            [x for x in df_S3_0.iloc[7:20, 0].tolist() if isinstance(x, int)],  # Range 1: A8:A23
            [x for x in df_S3_0.iloc[22:23, 0].tolist() if isinstance(x, int)],  # Range 2: A23
            [x for x in df_S3_0.iloc[25:44, 0].tolist() if isinstance(x, int)],  # Range 3: A26:A44
            [x for x in df_S3_0.iloc[47:55, 0].tolist() if isinstance(x, int)]   # Range 4: A48:A55
        ]
        col_ranges_map['Worksheet_S-3_Pts I-III'] = [
            [str(x) for x in df_S3_0.iloc[6, 2:10].tolist()],   # Range 1: C7:J7
            [str(x) for x in df_S3_0.iloc[21, 2:3].tolist()],   # Range 2: C22
            [str(x) for x in df_S3_0.iloc[24, 1:9].tolist()],   # Range 3: B25:I25
            [str(x) for x in df_S3_0.iloc[46, 8:9].tolist()]    # Range 4: I47
        ]   

# Read the Excel file for Worksheet_S-3_Pt IV
        df_S3_4 = pd.read_excel(xls, 'Worksheet_S-3_Pt IV', header=None)
        line_ranges_map['Worksheet_S-3_Pt IV'] = [
            [x for x in df_S3_4.iloc[6:24, 0].tolist() if isinstance(x, int)],  # Range 1: A7:A24
        ]
        col_ranges_map['Worksheet_S-3_Pt IV'] = [
            [str(x) for x in df_S3_4.iloc[5, 2:7].tolist()],   # Range 1: C6:G6
        ]

# Read the Excel file for Worksheet_S-3_Pt V
        df_S3_5 = pd.read_excel(xls, 'Worksheet_S-3_Pt V', header=None)
        line_ranges_map['Worksheet_S-3_Pt V'] = [
            [x for x in df_S3_5.iloc[7:36, 0].tolist() if isinstance(x, int)],  # Range 1: A8:A36
        ]
        col_ranges_map['Worksheet_S-3_Pt V'] = [
            [str(x) for x in df_S3_5.iloc[4, 3:8].tolist()],   # Range 1: D5:I5
        ]
                
        df_A = pd.read_excel(xls, 'Worksheet_A', header=None)
        line_ranges_map['Worksheet_A'] = [[x for x in df_A.iloc[4:46, 0].tolist() if isinstance(x, int)]]  #A5:A46
        col_ranges_map['Worksheet_A'] = [
            [str(x) for x in df_A.iloc[2, 1:12].tolist()]  #B3:L3
        ]
        
        df_B = pd.read_excel(xls, 'Worksheet_B', header=None)
        line_ranges_map['Worksheet_B'] = [
            [x for x in df_B.iloc[4:46, 0].tolist() if isinstance(x, int)],  # for the first range A5:A46
            [x for x in df_B.iloc[50:92, 0].tolist() if isinstance(x, int)]  # for the second range A51:A92
        ]
        col_ranges_map['Worksheet_B'] = [
            [str(x) for x in df_B.iloc[2, 2:9].tolist()], #C3:I3
            [str(x) for x in df_B.iloc[48, 2:9].tolist()]  #C49:I49
        ]  
        df_C = pd.read_excel(xls, 'Worksheet_C', header=None)
        line_ranges_map['Worksheet_C'] = [
            [x for x in df_C.iloc[7:17, 0].tolist() if isinstance(x, int)],  #A8:A17
            [x for x in df_C.iloc[23:27, 0].tolist() if isinstance(x, int)]   #A24:A27
        ]  
        col_ranges_map['Worksheet_C'] = [
            [str(x) for x in df_C.iloc[6, 7:12].tolist()],   #H7:L7
            [str(x) for x in df_C.iloc[22, 3:12].tolist()]   #D23:L23
        ]
    
        df_F = pd.read_excel(xls, 'Worksheet_F', header=None)
        line_ranges_map['Worksheet_F'] = [[x for x in df_F.iloc[5:62, 0].tolist() if isinstance(x, int)]]  #A6:A62
        col_ranges_map['Worksheet_F'] = [
            [str(x) for x in df_F.iloc[2, 4:5].tolist()]  #E3
        ]
    
        df_F1 = pd.read_excel(xls, 'Worksheet_F-1', header=None)
        line_ranges_map['Worksheet_F-1'] = [
            [x for x in df_F1.iloc[5:8, 0].tolist() if isinstance(x, int)], # A6:A8
            [x for x in df_F1.iloc[10:42, 0].tolist() if isinstance(x, int)] # A11:A42
        ]  # Two different ranges
        col_ranges_map['Worksheet_F-1'] = [
            [str(x) for x in df_F1.iloc[4, 2:6].tolist()],  # C5:F5
            [str(x) for x in df_F1.iloc[9, 1:6].tolist()]   # B10:F10
        ]       
    # Step 1: Loop through each target worksheet
    with pd.ExcelWriter(destination_path, engine='openpyxl', mode='a') as writer:
        #print("Entered Excel Writer block.")
        #initialize a "master" dictionary to hold line_to_row and col_to_column mappings for each range in each worksheet
        master_line_to_row = {}
        master_col_to_column = {}
        for worksheet_name in line_ranges_map.keys():
           # print(f"Working on worksheet: {worksheet_name}.")
            ws = writer.book[worksheet_name]
            # Initialize an empty mapping dictionary
            line_to_row = {}
            col_to_column = {}
            # Loop through the list of starting and ending points for line and column
            for idx, (start_row, end_row, start_col, end_col) in enumerate(zip(
                line_start_map[worksheet_name], line_end_map[worksheet_name], 
                col_start_map[worksheet_name], col_end_map[worksheet_name])):
                #print(f"Line to Row Mapping: {line_to_row}")  
                #print(f"Col to Column Mapping: {col_to_column}")
                # Create a unique key for each range within each worksheet
                unique_range_key = f"{worksheet_name}_{idx}"
                # Store line_to_row and col_to_column mapping for each unique range
                master_line_to_row[unique_range_key] = map_excel_rows_to_lines(ws, start_row, end_row)
                #print(f"Current line_to_row mapping for {unique_range_key}: {master_line_to_row[unique_range_key]}")
                # Fetch the row index to look for columns based on the current range
                col_row_index = col_row_map.get(worksheet_name, [3])[idx]  # default to 3 if not found
                master_col_to_column[unique_range_key] = map_excel_cols_to_columns(ws, start_col, end_col, col_row_index)
                #print(f"Current col_to_column mapping for {unique_range_key}: {master_col_to_column[unique_range_key]}")
                # Print the types of the column values stored within the master_col_to_column dictionary for the current unique_range_key
                #for col_key, col_value in master_col_to_column[unique_range_key].items():
                    #print(f"Column Key: {col_key}, Type: {type(col_key)}, Column Value: {col_value}, Type: {type(col_value)}")
            # Continue with the code that populates the Excel sheet
            df_filtered = populating_df[populating_df['WorksheetName'] == worksheet_name]        
            for _, row in df_filtered.iterrows():
                #print(f"DataFrame Row: {row}")
                line = row['ExcelRow']
                col = row['ExcelColumn']
                value = row['Value']
                #print(f"Type and value of line: {type(line)}, {line}") 
                #print(f"Type and value of col: {type(col)}, {col}")
                #print(f"Debug - About to look up col: {col} of type: {type(col)}")
                # Determine which unique range key to use 
                unique_range_key = determine_unique_range_key(line, col, line_ranges_map, col_ranges_map, worksheet_name)
                # Check if a unique_range_key was found
                if unique_range_key is None:
                    #print(f"Skipping: No unique range key found for Line {line}, Column {col}")
                    continue  # Skip to next iteration
                if unique_range_key:
                    excel_row = master_line_to_row.get(unique_range_key, {}).get(line)
                    excel_col = master_col_to_column.get(unique_range_key, {}).get(col)
                if excel_row is not None and excel_col is not None:
                    #print(f"Writing to Excel: Row {excel_row}, Column {excel_col}, Value {value}")
                    ws.cell(row=excel_row, column=excel_col, value=value)
                #else:
                    #print(f"Skipping: Excel Row or Excel Col is None. Excel Row: {excel_row}, Excel Col: {excel_col}")
    # Save the workbook again after adding the sheets, populating, and formatting
        destination_wb.save(destination_path)



    


















